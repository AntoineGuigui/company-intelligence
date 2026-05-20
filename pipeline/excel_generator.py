"""
pipeline/excel_generator.py — Write structured data into DataBase.xlsm format.

Writes to the same Excel schema used by the Company Profile Generator,
so generated data feeds directly into the PPTX generation pipeline.

Sheets:
  - DataBase: main company data (1 row per company)
  - Revenue: revenue by FY (1 row per company, year columns)
  - EBIT: EBIT by FY
  - Net Profit: net profit by FY
  - EBIT Margin: EBIT margin % by FY
  - Net Profit Margin: net profit margin % by FY
"""

import logging
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

# ── DataBase sheet columns (must match company_profile_generator.py) ──
DATABASE_COLUMNS = [
    "Company Name",
    "Country",
    "Field",
    "Activity",
    "Locations",
    "Founded",
    "N° employees",
    "Key people",
    "Type Ownership",
    "Confidence Index",
    "Business Overview",
    "Business relationships",
    "Capability",
    "Notes",
]

# ── Financial sheet names ──
FINANCIAL_SHEETS = ["Revenue", "EBIT", "Net Profit", "EBIT Margin", "Net Profit Margin"]

# ── Mapping from extractor JSON financials keys to sheet names ──
FIN_KEY_MAP = {
    "Revenue": "revenues",
    "EBIT": "ebit",
    "Net Profit": "net_profit",
    "EBIT Margin": "ebit_margin",
    "Net Profit Margin": "net_profit_margin",
}


def _ensure_workbook(path: str) -> Workbook:
    """Load existing workbook or create a new one with correct structure."""
    p = Path(path)

    if p.exists():
        wb = load_workbook(p, keep_vba=p.suffix == ".xlsm")
        return wb

    # Create new workbook with all required sheets
    wb = Workbook()

    # DataBase sheet
    ws = wb.active
    ws.title = "DataBase"
    for col_idx, col_name in enumerate(DATABASE_COLUMNS, start=1):
        ws.cell(row=1, column=col_idx, value=col_name)

    # Financial sheets
    for sheet_name in FINANCIAL_SHEETS:
        ws_fin = wb.create_sheet(title=sheet_name)
        ws_fin.cell(row=1, column=1, value="Company")

    return wb


def _map_to_database_row(data: dict) -> dict:
    """Map extractor output to DataBase sheet columns."""
    return {
        "Company Name": data.get("company_name", ""),
        "Country": data.get("country", ""),
        "Field": data.get("field", ""),
        "Activity": data.get("activity", ""),
        "Locations": data.get("locations", ""),
        "Founded": data.get("founded", ""),
        "N° employees": data.get("employees", ""),
        "Key people": data.get("key_people", ""),
        "Type Ownership": data.get("type_ownership", ""),
        "Confidence Index": str(data.get("confidence_index", "1")),
        "Business Overview": data.get("business_overview", ""),
        "Business relationships": data.get("business_relationships", ""),
        "Capability": data.get("capability", ""),
        "Notes": data.get("notes", ""),
    }


def _find_or_add_company_row(ws, company_name: str) -> int:
    """Find existing row for company or return next empty row."""
    for row_idx in range(2, ws.max_row + 1):
        cell_val = ws.cell(row=row_idx, column=1).value
        if cell_val and str(cell_val).strip().lower() == company_name.strip().lower():
            return row_idx
    return ws.max_row + 1


def _fy_to_year(fy_str: str) -> int:
    """Convert 'FY24' to 2024."""
    num = int(fy_str.replace("FY", ""))
    return 2000 + num if num < 100 else num


def _ensure_year_column(ws, year: int) -> int:
    """Find or create a column for the given year. Returns column index."""
    # Scan existing headers
    for col_idx in range(2, ws.max_column + 1):
        header = ws.cell(row=1, column=col_idx).value
        if header is not None:
            try:
                if int(header) == year:
                    return col_idx
            except (ValueError, TypeError):
                pass

    # Add new column
    col_idx = max(ws.max_column + 1, 2)
    ws.cell(row=1, column=col_idx, value=year)
    return col_idx


def generate_excel(data: dict, excel_path: str) -> str:
    """
    Write structured company data into a DataBase.xlsm-compatible Excel file.

    If the company already exists, its row is updated.
    If not, a new row is appended.

    Args:
        data: Structured dict from extractor.extract().
        excel_path: Path to the Excel file.

    Returns:
        Path to the saved file.
    """
    wb = _ensure_workbook(excel_path)

    company_name = data.get("company_name", "Unknown")

    # ── 1. Write to DataBase sheet ──
    ws_db = wb["DataBase"]
    row_data = _map_to_database_row(data)
    row_idx = _find_or_add_company_row(ws_db, company_name)

    for col_idx, col_name in enumerate(DATABASE_COLUMNS, start=1):
        value = row_data.get(col_name, "")
        ws_db.cell(row=row_idx, column=col_idx, value=value)

    logger.info(f"Wrote {company_name} to DataBase row {row_idx}")

    # ── 2. Write financial data ──
    financials = data.get("financials", {})

    for sheet_name, json_key in FIN_KEY_MAP.items():
        if sheet_name not in wb.sheetnames:
            ws_fin = wb.create_sheet(title=sheet_name)
            ws_fin.cell(row=1, column=1, value="Company")
        else:
            ws_fin = wb[sheet_name]

        # Find or add company row in this financial sheet
        fin_row = _find_or_add_company_row(ws_fin, company_name)
        ws_fin.cell(row=fin_row, column=1, value=company_name)

        # Write each FY value
        for fy_str, fy_data in financials.items():
            if not isinstance(fy_data, dict):
                continue
            value = fy_data.get(json_key)
            if value is None:
                continue

            try:
                year = _fy_to_year(fy_str)
                col_idx = _ensure_year_column(ws_fin, year)
                ws_fin.cell(row=fin_row, column=col_idx, value=float(value))
            except (ValueError, TypeError) as e:
                logger.warning(f"Skipping {sheet_name}/{fy_str}: {e}")

    # ── 3. Save ──
    p = Path(excel_path)
    if p.suffix == ".xlsm":
        wb.save(excel_path)
    else:
        wb.save(excel_path)

    logger.info(f"Saved Excel: {excel_path}")
    return excel_path
